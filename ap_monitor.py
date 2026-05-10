import smtplib
import warnings
import io
import re
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from ping3 import ping
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────
#  CONFIGURATION
#  Set these as environment variables or
#  fill them in a local config file that
#  is listed in .gitignore.
# ─────────────────────────────────────────
OMADA_URL  = os.environ.get("OMADA_URL",  "https://<controller-ip>:<port>")
OMADA_USER = os.environ.get("OMADA_USER", "admin")
OMADA_PASS = os.environ.get("OMADA_PASS", "")
SITE_NAME  = os.environ.get("OMADA_SITE", "MySite")
COMPANY    = os.environ.get("COMPANY",    "MyCompany")

ZOHO_SMTP  = os.environ.get("ZOHO_SMTP",  "smtp.zoho.com")
ZOHO_PORT  = int(os.environ.get("ZOHO_PORT", 587))
ZOHO_USER  = os.environ.get("ZOHO_USER",  "you@example.com")
ZOHO_PASS  = os.environ.get("ZOHO_PASS",  "")
MAIL_TO    = os.environ.get("MAIL_TO",    "you@example.com").split(",")

# ─────────────────────────────────────────
#  STATIC MAC -> LOCATION + IP MAPPING
#
#  Format:
#    "aa-bb-cc-dd-ee-ff": {"ip": "192.168.x.x", "location": "Description"},
#
#  Tip: load this from a JSON file that is
#  listed in .gitignore to keep site data
#  out of version control.
# ─────────────────────────────────────────
MAC_INFO = {
    # Example entries — replace with your own:
    # "aa-bb-cc-dd-ee-01": {"ip": "192.168.1.10", "location": "Floor 1 - Room A"},
    # "aa-bb-cc-dd-ee-02": {"ip": "192.168.1.11", "location": "Floor 2 - Room B"},
}

# Alternatively, load from an external JSON file:
# import json
# MAC_INFO_PATH = os.environ.get("MAC_INFO_PATH", "mac_info.json")
# if os.path.exists(MAC_INFO_PATH):
#     with open(MAC_INFO_PATH) as f:
#         MAC_INFO = json.load(f)

# ─────────────────────────────────────────

session    = requests.Session()
session.verify = False
OMADA_CID  = None
CSRF_TOKEN = None


def _headers():
    h = {"Content-Type": "application/json", "Referer": f"{OMADA_URL}/"}
    if CSRF_TOKEN:
        h["Csrf-Token"] = CSRF_TOKEN
    return h


def omada_login():
    global OMADA_CID, CSRF_TOKEN
    print("Fetching controller info...")
    r         = session.get(f"{OMADA_URL}/api/info")
    result    = r.json().get("result", {})
    OMADA_CID = result.get("omadacId", "")
    print(f"Controller: v{result.get('controllerVer')}  |  ID: {OMADA_CID}")

    print("Logging in...")
    r    = session.post(f"{OMADA_URL}/{OMADA_CID}/api/v2/login",
                        json={"username": OMADA_USER, "password": OMADA_PASS},
                        headers={"Content-Type": "application/json"})
    data = r.json()
    if data.get("errorCode") != 0:
        raise Exception(f"Login failed ({data.get('errorCode')}): {data.get('msg')}")
    CSRF_TOKEN = data.get("result", {}).get("token", "")
    print(f"Login successful! Token: {CSRF_TOKEN[:12]}...")


def get_site_id():
    print(f"Resolving site: '{SITE_NAME}'...")
    r     = session.get(f"{OMADA_URL}/{OMADA_CID}/api/v2/user/sites",
                        headers=_headers(), params={"pageSize": 100, "page": 1})
    sites = r.json().get("result", {}).get("data", [])
    for site in sites:
        if site.get("name", "").lower() == SITE_NAME.lower():
            sid = site.get("id") or site.get("siteId")
            print(f"Site: '{SITE_NAME}' -> {sid}")
            return sid
    if sites:
        sid = sites[0].get("id") or sites[0].get("siteId")
        print(f"Using first site: {sites[0].get('name')} -> {sid}")
        return sid
    raise Exception("No sites found.")


def fetch_aps(site_id):
    print("Fetching APs...")
    all_aps, page, page_size = [], 1, 100
    while True:
        r      = session.get(f"{OMADA_URL}/{OMADA_CID}/api/v2/sites/{site_id}/devices",
                             headers=_headers(),
                             params={"pageSize": page_size, "page": page})
        data   = r.json()
        result = data.get("result", [])
        if isinstance(result, list):
            devices = result
            total   = len(result)
        else:
            devices = result.get("data", [])
            total   = result.get("totalRows", len(devices))
        all_aps.extend(devices)
        print(f"  Page {page} -> {len(devices)} devices")
        if page * page_size >= total:
            break
        page += 1
    ap_list = [d for d in all_aps if str(d.get("type", "")).lower() in ("0", "ap")]
    if not ap_list:
        ap_list = all_aps
    print(f"Found {len(ap_list)} Access Point(s)")
    return ap_list


def ping_check(ip):
    if not ip or ip in ("NA", "None", "", "0.0.0.0"):
        return False, "No IP"
    try:
        result = ping(ip, timeout=3)
        return (True, "UP") if result is not None else (False, "DOWN")
    except Exception:
        return False, "Error"


def to_str(val, default="NA"):
    if val is None or val == "":
        return default
    if isinstance(val, (dict, list)):
        return default
    return str(val).strip() or default


def normalise_mac(mac):
    if not mac:
        return ""
    clean = re.sub(r"[:\-\.]", "", str(mac)).lower()
    if len(clean) == 12:
        return "-".join(clean[i:i+2] for i in range(0, 12, 2))
    return str(mac).lower()


def format_uptime(raw):
    """
    Convert seconds (int/float) to '208 days 5 h 58 m 45 s' format.
    If the controller already returns a string, return it as-is.
    """
    if raw is None or raw == "" or raw == 0 or raw == "0":
        return "NA"

    if isinstance(raw, (int, float)):
        seconds = int(raw)
        if seconds <= 0:
            return "NA"
        days    = seconds // 86400
        hours   = (seconds % 86400) // 3600
        minutes = (seconds % 3600) // 60
        secs    = seconds % 60
        parts   = []
        if days:    parts.append(f"{days} days")
        if hours:   parts.append(f"{hours} h")
        if minutes: parts.append(f"{minutes} m")
        if secs:    parts.append(f"{secs} s")
        return " ".join(parts) if parts else "< 1 s"

    s = str(raw).strip()
    try:
        seconds = int(float(s))
        if seconds > 0:
            days    = seconds // 86400
            hours   = (seconds % 86400) // 3600
            minutes = (seconds % 3600) // 60
            secs    = seconds % 60
            parts   = []
            if days:    parts.append(f"{days} days")
            if hours:   parts.append(f"{hours} h")
            if minutes: parts.append(f"{minutes} m")
            if secs:    parts.append(f"{secs} s")
            return " ".join(parts) if parts else "< 1 s"
    except Exception:
        pass

    return s


def process_aps(raw_aps):
    enriched = []
    for ap in raw_aps:
        is_connected = ap.get("statusCategory", 0) == 1
        raw_mac      = to_str(ap.get("mac"), "NA")
        norm_mac     = normalise_mac(raw_mac)

        live_ip   = to_str(ap.get("ip") or ap.get("lanIp"), "NA")
        mapped_ip = MAC_INFO.get(norm_mac, {}).get("ip", "")
        ip        = mapped_ip if mapped_ip else live_ip

        ok, ping_str = ping_check(ip)

        ch2g = to_str(ap.get("wp2g", {}).get("actualChannel", ""), "")
        ch5g = to_str(ap.get("wp5g", {}).get("actualChannel", ""), "")
        ch2g = ch2g.split("/")[0].strip() if ch2g else ""
        ch5g = ch5g.split("/")[0].strip() if ch5g else ""
        channels = []
        if ch2g: channels.append(f"{ch2g}(2.4G)")
        if ch5g: channels.append(f"{ch5g}(5G)")
        channel = ", ".join(channels) if channels else "NA"

        raw_uptime = (
            ap.get("uptime") or ap.get("uptimeLong") or
            ap.get("uptimeSeconds") or ap.get("connectedTime") or
            ap.get("onlineTime") or 0
        )
        uptime_display = format_uptime(raw_uptime)

        mapped_location = MAC_INFO.get(norm_mac, {}).get("location", "")
        if mapped_location:
            location = mapped_location
        else:
            location = to_str(
                ap.get("location") or ap.get("zone") or ap.get("area") or ap.get("site"),
                default=to_str(ap.get("name") or raw_mac, "NA")
            )

        enriched.append({
            "device_name": to_str(ap.get("name") or raw_mac, "Unknown"),
            "ip":          ip,
            "mac":         raw_mac,
            "status":      "Connected" if is_connected else "Offline",
            "connected":   is_connected,
            "ping_ok":     ok,
            "ping_str":    ping_str,
            "model":       to_str(ap.get("compoundModel") or ap.get("showModel") or ap.get("model"), "NA"),
            "version":     to_str(ap.get("firmwareVersion") or ap.get("version"), "NA"),
            "uptime":      uptime_display,
            "clients":     str(int(ap.get("clientNum", 0))),
            "channel":     channel,
            "cpu":         f"{ap.get('cpuUtil', 0)}%",
            "mem":         f"{ap.get('memUtil', 0)}%",
            "label":       to_str(ap.get("name") or raw_mac, "NA"),
            "location":    location,
        })
    return enriched


# ─────────────────────────────────────────
#  BUILD EXCEL
# ─────────────────────────────────────────
def build_excel(enriched, connected, offline, total, now):
    health_pct = round((connected / total) * 100) if total > 0 else 0
    wb = Workbook()
    ws = wb.active
    ws.title = "AP Status"

    NAVY     = "1F3864"
    WHITE    = "FFFFFF"
    GRAY_LT  = "F5F5F5"
    GRAY_MID = "CCCCCC"
    GREEN_BG = "E2EFDA"
    GREEN_FG = "375623"
    RED_BG   = "FCE4D6"
    RED_FG   = "843C0C"

    def fill(c):  return PatternFill("solid", fgColor=c)
    def ctr():    return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    def bd():
        s = Side(style="thin", color=GRAY_MID)
        return Border(left=s, right=s, top=s, bottom=s)

    ws.merge_cells("A1:O1")
    ws["A1"] = f"{COMPANY} - Access Point Report  |  {datetime.now().strftime('%d/%m/%Y')}  |  Site: {SITE_NAME.upper()}"
    ws["A1"].font      = Font(name="Calibri", size=13, bold=True, color=WHITE)
    ws["A1"].fill      = fill(NAVY)
    ws["A1"].alignment = lft()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:O2")
    ws["A2"] = f"Generated: {now}   |   Total: {total}   |   Connected: {connected}   |   Offline: {offline}   |   Health: {health_pct}%"
    ws["A2"].font      = Font(name="Calibri", size=9, color=WHITE)
    ws["A2"].fill      = fill(NAVY)
    ws["A2"].alignment = lft()
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    headers    = ["#", "LOCATION", "IP ADDRESS", "MAC", "STATUS", "PING",
                  "MODEL", "FIRMWARE", "UPTIME", "CLIENTS", "CHANNEL", "CPU", "MEM"]
    col_widths = [5, 42, 16, 20, 13, 10, 22, 28, 22, 9, 20, 8, 8]
    for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=4, column=ci, value=hdr)
        cell.font      = Font(name="Calibri", size=9, bold=True, color=WHITE)
        cell.fill      = fill(NAVY)
        cell.alignment = ctr()
        cell.border    = bd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 18

    for ri, ap in enumerate(enriched, start=5):
        row_num = ri - 4
        row_bg  = WHITE if row_num % 2 != 0 else GRAY_LT
        vals    = [row_num, ap["location"], ap["ip"], ap["mac"], ap["status"],
                   ap["ping_str"], ap["model"], ap["version"], ap["uptime"],
                   ap["clients"], ap["channel"], ap["cpu"], ap["mem"]]
        for ci, val in enumerate(vals, 1):
            if isinstance(val, (dict, list)):
                val = "NA"
            cell           = ws.cell(row=ri, column=ci, value=val)
            cell.border    = bd()
            cell.alignment = ctr() if ci in [1, 5, 6, 10, 12, 13] else lft()
            cell.font      = Font(name="Calibri", size=9, color="000000")
            cell.fill      = fill(row_bg)
            if ci == 5:
                is_conn = ap["connected"]
                cell.font = Font(name="Calibri", size=9, bold=True,
                                 color=GREEN_FG if is_conn else RED_FG)
                cell.fill = fill(GREEN_BG if is_conn else RED_BG)
            if ci == 6:
                cell.font = Font(name="Calibri", size=9, bold=True,
                                 color=GREEN_FG if ap["ping_ok"] else RED_FG)
                cell.fill = fill(GREEN_BG if ap["ping_ok"] else RED_BG)
        ws.row_dimensions[ri].height = 16

    fr = len(enriched) + 6
    ws.merge_cells(f"A{fr}:M{fr}")
    ws[f"A{fr}"] = f"{COMPANY} - Auto-generated report from Omada Controller"
    ws[f"A{fr}"].font      = Font(name="Calibri", size=8, color="888888")
    ws[f"A{fr}"].alignment = lft()
    ws.row_dimensions[fr].height = 14
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────
#  BUILD HTML EMAIL  (Outlook-safe)
# ─────────────────────────────────────────
def build_html_email(enriched, connected, offline, total, now):
    health_pct = round((connected / total) * 100) if total > 0 else 0

    if offline > 0:
        alert_bg  = "#FFF3CD"; alert_bdr = "#FFC107"; alert_txt = "#7D4E00"
        alert_msg = f"<b>Warning:</b> {offline} access point(s) are currently OFFLINE."
    else:
        alert_bg  = "#D4EDDA"; alert_bdr = "#28A745"; alert_txt = "#155724"
        alert_msg = "<b>All Clear:</b> All access points are online and responding."

    rows_html = ""
    for i, ap in enumerate(enriched):
        row_bg = "#FFFFFF" if i % 2 == 0 else "#F9F9F9"

        if ap["connected"]:
            status_color = "#155724"; status_bg = "#D4EDDA"; status_text = "Connected"
        else:
            status_color = "#721C24"; status_bg = "#F8D7DA"; status_text = "Offline"

        uptime_val   = ap["uptime"]   if ap["uptime"]   not in ("NA", "", None) else "&mdash;"
        location_val = ap["location"] if ap["location"] not in ("NA", "", None) else "&mdash;"
        ip_val       = ap["ip"]

        rows_html += f"""
<tr>
  <td style="padding:8px 12px;border:1px solid #DDDDDD;background-color:{row_bg};font-family:Arial,sans-serif;font-size:13px;color:#222222;word-wrap:break-word;">{location_val}</td>
  <td style="padding:8px 12px;border:1px solid #DDDDDD;background-color:{row_bg};font-family:Courier New,monospace;font-size:12px;color:#1A56A0;white-space:nowrap;">{ip_val}</td>
  <td style="padding:8px 12px;border:1px solid #DDDDDD;background-color:{row_bg};font-family:Arial,sans-serif;font-size:12px;color:#555555;text-align:center;white-space:nowrap;">{uptime_val}</td>
  <td style="padding:8px 12px;border:1px solid #DDDDDD;background-color:{status_bg};font-family:Arial,sans-serif;font-size:12px;font-weight:bold;color:{status_color};text-align:center;white-space:nowrap;">{status_text}</td>
</tr>"""

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background-color:#EEEEEE;">
<table width="100%" cellpadding="0" cellspacing="0" style="background-color:#EEEEEE;">
<tr><td align="center" style="padding:20px;">
<table width="700" cellpadding="0" cellspacing="0" style="background-color:#FFFFFF;border:1px solid #CCCCCC;">

  <tr>
    <td style="background-color:#1F3864;padding:18px 24px;">
      <p style="margin:0;font-family:Arial,sans-serif;font-size:11px;color:#AAAAAA;letter-spacing:1px;text-transform:uppercase;">Network Monitoring Report</p>
      <p style="margin:4px 0 0 0;font-family:Arial,sans-serif;font-size:20px;font-weight:bold;color:#FFFFFF;">{COMPANY}</p>
      <p style="margin:4px 0 0 0;font-family:Arial,sans-serif;font-size:12px;color:#BBDDFF;">Access Point Status &nbsp;|&nbsp; Site: {SITE_NAME.upper()} &nbsp;|&nbsp; {datetime.now().strftime('%d %B %Y')}</p>
    </td>
  </tr>

  <tr>
    <td style="background-color:#F0F4F8;padding:8px 24px;border-bottom:1px solid #CCCCCC;">
      <p style="margin:0;font-family:Arial,sans-serif;font-size:11px;color:#666666;">Generated: {now}</p>
    </td>
  </tr>

  <tr>
    <td style="padding:16px 24px;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td width="25%" align="center" style="padding:4px;">
            <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #CCCCCC;">
              <tr><td style="padding:14px 8px;text-align:center;">
                <p style="margin:0;font-family:Arial,sans-serif;font-size:24px;font-weight:bold;color:#1F3864;">{total}</p>
                <p style="margin:4px 0 0 0;font-family:Arial,sans-serif;font-size:10px;color:#666666;text-transform:uppercase;letter-spacing:1px;">Total APs</p>
              </td></tr>
            </table>
          </td>
          <td width="25%" align="center" style="padding:4px;">
            <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #28A745;">
              <tr><td style="padding:14px 8px;text-align:center;background-color:#D4EDDA;">
                <p style="margin:0;font-family:Arial,sans-serif;font-size:24px;font-weight:bold;color:#155724;">{connected}</p>
                <p style="margin:4px 0 0 0;font-family:Arial,sans-serif;font-size:10px;color:#155724;text-transform:uppercase;letter-spacing:1px;">Connected</p>
              </td></tr>
            </table>
          </td>
          <td width="25%" align="center" style="padding:4px;">
            <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid {'#DC3545' if offline > 0 else '#CCCCCC'};">
              <tr><td style="padding:14px 8px;text-align:center;background-color:{'#F8D7DA' if offline > 0 else '#FFFFFF'};">
                <p style="margin:0;font-family:Arial,sans-serif;font-size:24px;font-weight:bold;color:{'#721C24' if offline > 0 else '#888888'};">{offline}</p>
                <p style="margin:4px 0 0 0;font-family:Arial,sans-serif;font-size:10px;color:{'#721C24' if offline > 0 else '#888888'};text-transform:uppercase;letter-spacing:1px;">Offline</p>
              </td></tr>
            </table>
          </td>
        </tr>
      </table>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>"""
    return html
